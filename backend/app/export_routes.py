from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from datetime import datetime, timedelta
from typing import Optional
import io

from app.database import get_db
from app.models import Contact, Message, User, Tag, contact_tags
from app.auth import get_current_user

router = APIRouter(prefix="/export", tags=["export"])


def style_header(ws, headers, col_widths):
    """Aplica estilo no cabeçalho da planilha"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        bottom=Side(style="thin", color="E5E7EB")
    )
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = width
    
    # Ajustar larguras por letra
    from openpyxl.utils import get_column_letter
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30


def style_data_row(ws, row_idx, num_cols):
    """Aplica estilo nas linhas de dados"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data_font = Font(name="Arial", size=10)
    alt_fill = PatternFill(start_color="F8F9FB", end_color="F8F9FB", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="F3F4F6"))

    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
        if row_idx % 2 == 0:
            cell.fill = alt_fill


STATUS_LABELS = {
    "novo": "Novo Lead",
    "em_contato": "Em Contato",
    "qualificado": "Qualificado",
    "negociando": "Em Matrícula",
    "convertido": "Matriculado",
    "perdido": "Perdido",
}


@router.get("/contacts")
async def export_contacts(
    channel_id: Optional[int] = None,
    status: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Exporta relatório de contatos em Excel"""
    import openpyxl

    # Buscar contatos
    filters = []
    if channel_id:
        filters.append(Contact.channel_id == channel_id)
    if status:
        filters.append(Contact.lead_status == status)

    result = await db.execute(
        select(Contact).where(*filters).order_by(Contact.created_at.desc())
    )
    contacts = result.scalars().all()

    # Buscar usuários para nome do atribuído
    users_q = await db.execute(select(User))
    users_map = {u.id: u.name for u in users_q.scalars().all()}

    # Buscar tags por contato
    tags_q = await db.execute(
        select(contact_tags.c.contact_wa_id, Tag.name)
        .join(Tag, Tag.id == contact_tags.c.tag_id)
    )
    tags_map: dict = {}
    for row in tags_q.all():
        tags_map.setdefault(row[0], []).append(row[1])

    # Buscar contagem de mensagens por contato
    msgs_q = await db.execute(
        select(
            Message.contact_wa_id,
            func.count(Message.id).label("total"),
            func.count(Message.id).filter(Message.direction == "inbound").label("inbound"),
            func.count(Message.id).filter(Message.direction == "outbound").label("outbound"),
        ).group_by(Message.contact_wa_id)
    )
    msgs_map = {row[0]: {"total": row[1], "in": row[2], "out": row[3]} for row in msgs_q.all()}

    # Criar Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contatos"

    headers = [
        "Nome", "Telefone", "Status", "Atribuído a",
        "Tags", "Mensagens (total)", "Recebidas", "Enviadas",
        "Notas", "Criado em"
    ]
    col_widths = [25, 18, 16, 20, 25, 16, 14, 14, 35, 18]
    style_header(ws, headers, col_widths)

    for idx, c in enumerate(contacts, 2):
        msg_data = msgs_map.get(c.wa_id, {"total": 0, "in": 0, "out": 0})
        contact_tags_list = tags_map.get(c.wa_id, [])

        ws.cell(row=idx, column=1, value=c.name or "Sem nome")
        ws.cell(row=idx, column=2, value=c.wa_id)
        ws.cell(row=idx, column=3, value=STATUS_LABELS.get(c.lead_status or "novo", c.lead_status))
        ws.cell(row=idx, column=4, value=users_map.get(c.assigned_to, "—") if c.assigned_to else "—")
        ws.cell(row=idx, column=5, value=", ".join(contact_tags_list) if contact_tags_list else "—")
        ws.cell(row=idx, column=6, value=msg_data["total"])
        ws.cell(row=idx, column=7, value=msg_data["in"])
        ws.cell(row=idx, column=8, value=msg_data["out"])
        ws.cell(row=idx, column=9, value=(c.notes or "")[:100])
        ws.cell(row=idx, column=10, value=c.created_at.strftime("%d/%m/%Y %H:%M") if c.created_at else "—")
        style_data_row(ws, idx, len(headers))

    # Resumo
    ws_resumo = wb.create_sheet("Resumo")
    resumo_headers = ["Métrica", "Valor"]
    style_header(ws_resumo, resumo_headers, [30, 20])

    status_counts: dict = {}
    for c in contacts:
        s = c.lead_status or "novo"
        status_counts[s] = status_counts.get(s, 0) + 1

    resumo_data = [
        ("Total de Contatos", len(contacts)),
        ("Com atribuição", sum(1 for c in contacts if c.assigned_to)),
        ("Sem atribuição", sum(1 for c in contacts if not c.assigned_to)),
        ("", ""),
        ("--- Por Status ---", ""),
    ]
    for status_key, label in STATUS_LABELS.items():
        resumo_data.append((label, status_counts.get(status_key, 0)))

    for idx, (metrica, valor) in enumerate(resumo_data, 2):
        ws_resumo.cell(row=idx, column=1, value=metrica)
        ws_resumo.cell(row=idx, column=2, value=valor)
        style_data_row(ws_resumo, idx, 2)

    # Salvar em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"relatorio_contatos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/pipeline")
async def export_pipeline(
    channel_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Exporta relatório do funil/pipeline em Excel"""
    import openpyxl

    filters = []
    if channel_id:
        filters.append(Contact.channel_id == channel_id)

    result = await db.execute(
        select(Contact).where(*filters).order_by(Contact.lead_status, Contact.created_at.desc())
    )
    contacts = result.scalars().all()

    users_q = await db.execute(select(User))
    users_map = {u.id: u.name for u in users_q.scalars().all()}

    wb = openpyxl.Workbook()

    # Aba por status
    for status_key, label in STATUS_LABELS.items():
        status_contacts = [c for c in contacts if (c.lead_status or "novo") == status_key]

        ws = wb.create_sheet(title=label[:31])
        headers = ["Nome", "Telefone", "Atribuído a", "Notas", "Criado em"]
        col_widths = [25, 18, 20, 40, 18]
        style_header(ws, headers, col_widths)

        for idx, c in enumerate(status_contacts, 2):
            ws.cell(row=idx, column=1, value=c.name or "Sem nome")
            ws.cell(row=idx, column=2, value=c.wa_id)
            ws.cell(row=idx, column=3, value=users_map.get(c.assigned_to, "—") if c.assigned_to else "—")
            ws.cell(row=idx, column=4, value=(c.notes or "")[:100])
            ws.cell(row=idx, column=5, value=c.created_at.strftime("%d/%m/%Y %H:%M") if c.created_at else "—")
            style_data_row(ws, idx, len(headers))

    # Resumo geral
    ws_resumo = wb.active
    ws_resumo.title = "Resumo do Funil"
    headers = ["Etapa", "Quantidade", "% do Total"]
    col_widths = [22, 16, 16]
    style_header(ws_resumo, headers, col_widths)

    total = len(contacts) or 1
    row = 2
    for status_key, label in STATUS_LABELS.items():
        count = sum(1 for c in contacts if (c.lead_status or "novo") == status_key)
        pct = round(count / total * 100, 1)
        ws_resumo.cell(row=row, column=1, value=label)
        ws_resumo.cell(row=row, column=2, value=count)
        ws_resumo.cell(row=row, column=3, value=f"{pct}%")
        style_data_row(ws_resumo, row, 3)
        row += 1

    ws_resumo.cell(row=row + 1, column=1, value="TOTAL")
    ws_resumo.cell(row=row + 1, column=2, value=len(contacts))
    from openpyxl.styles import Font
    ws_resumo.cell(row=row + 1, column=1).font = Font(name="Arial", size=11, bold=True)
    ws_resumo.cell(row=row + 1, column=2).font = Font(name="Arial", size=11, bold=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"relatorio_pipeline_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/messages")
async def export_messages(
    days: int = Query(default=7, ge=1, le=90),
    channel_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Exporta relatório de mensagens dos últimos N dias"""
    import openpyxl

    since = datetime.now() - timedelta(days=days)

    filters = [Message.timestamp >= since]
    if channel_id:
        filters.append(Message.channel_id == channel_id)

    result = await db.execute(
        select(Message).where(*filters).order_by(Message.timestamp.desc()).limit(5000)
    )
    messages = result.scalars().all()

    # Nomes dos contatos
    contacts_q = await db.execute(select(Contact))
    contacts_map = {c.wa_id: c.name for c in contacts_q.scalars().all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mensagens"

    headers = ["Data/Hora", "Contato", "Telefone", "Direção", "Tipo", "Conteúdo"]
    col_widths = [20, 25, 18, 14, 14, 60]
    style_header(ws, headers, col_widths)

    for idx, m in enumerate(messages, 2):
        content = m.content or ""
        if content.startswith("media:"):
            content = f"[{m.message_type or 'mídia'}]"

        ws.cell(row=idx, column=1, value=m.timestamp.strftime("%d/%m/%Y %H:%M") if m.timestamp else "—")
        ws.cell(row=idx, column=2, value=contacts_map.get(m.contact_wa_id, "Desconhecido"))
        ws.cell(row=idx, column=3, value=m.contact_wa_id)
        ws.cell(row=idx, column=4, value="Recebida" if m.direction == "inbound" else "Enviada")
        ws.cell(row=idx, column=5, value=m.message_type or "text")
        ws.cell(row=idx, column=6, value=content[:200])
        style_data_row(ws, idx, len(headers))

    # Resumo
    ws_resumo = wb.create_sheet("Resumo")
    resumo_headers = ["Métrica", "Valor"]
    style_header(ws_resumo, resumo_headers, [30, 20])

    inbound = sum(1 for m in messages if m.direction == "inbound")
    outbound = sum(1 for m in messages if m.direction == "outbound")

    # Msgs por dia
    daily: dict = {}
    for m in messages:
        if m.timestamp:
            day = m.timestamp.strftime("%d/%m/%Y")
            daily[day] = daily.get(day, 0) + 1

    resumo_data = [
        ("Período", f"Últimos {days} dias"),
        ("Total de Mensagens", len(messages)),
        ("Recebidas", inbound),
        ("Enviadas", outbound),
        ("Contatos únicos", len(set(m.contact_wa_id for m in messages))),
        ("Média por dia", round(len(messages) / max(days, 1), 1)),
    ]

    for idx, (metrica, valor) in enumerate(resumo_data, 2):
        ws_resumo.cell(row=idx, column=1, value=metrica)
        ws_resumo.cell(row=idx, column=2, value=valor)
        style_data_row(ws_resumo, idx, 2)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"relatorio_mensagens_{days}d_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )